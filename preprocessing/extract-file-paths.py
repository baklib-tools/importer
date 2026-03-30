#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
文件路径分类导出脚本

功能说明：
    读取包含文件路径的文本文件，根据文件扩展名自动分类，将图片、视频、音频、Office文件
    的路径分别导出到独立的文本文件中。

支持的文件类型：
    - 图片：jpg, jpeg, png, gif, webp, tiff, raw, svg
    - 视频：mp4, mov, m4v, swf, flv, avi, wmv, webm
    - 音频：mp3, wav, midi, wma, flac, ape, cda, aac
    - Office：doc, docx, xls, xlsx, ppt, pptx, pdf, txt, md 等

使用方法：
    # 基本用法（输出目录自动使用输入文件名）
    python3 preprocessing/extract-file-paths.py <输入文件路径>

    # 指定输出目录
    python3 preprocessing/extract-file-paths.py <输入文件路径> -o <输出目录>

    # 生成拆分文件（每N个记录一个文件）
    python3 preprocessing/extract-file-paths.py <输入文件路径> --split <N>

    # 输出为Excel格式
    python3 preprocessing/extract-file-paths.py <输入文件路径> --format excel

    # 查看帮助
    python3 preprocessing/extract-file-paths.py --help

参数说明：
    input_file    必需，输入文件路径（包含文件路径列表的文本文件）
    -o, --output  可选，输出目录路径（默认使用输入文件名作为目录名）
    -s, --split   可选，拆分文件大小（每个文件的最大记录数），如果指定此参数，会在汇总文件基础上生成拆分文件
    -f, --format  可选，拆分文件的输出格式：txt（文本文件，默认）或 excel（Excel文件）。注意：汇总文件始终为 txt 格式

输出文件：
    脚本会在输出目录中生成以下文件（文件名包含记录数）：
    - 图片文件路径-共N个.txt    - 包含所有图片文件路径（汇总文件，始终为txt格式）
    - 视频文件路径-共N个.txt    - 包含所有视频文件路径（汇总文件，始终为txt格式）
    - 音频文件路径-共N个.txt    - 包含所有音频文件路径（汇总文件，始终为txt格式）
    - Office文件路径-共N个.txt  - 包含所有Office和文本文件路径（汇总文件，始终为txt格式）
    （N为实际文件数量，如果某类文件数量为0，则不会生成对应文件）

    如果指定了 --split 参数，还会生成拆分文件：
    - 图片文件路径-第1-10000个.{txt|xlsx}    - 第1到10000个记录
    - 图片文件路径-第10001-20000个.{txt|xlsx} - 第10001到20000个记录
    （如果文件数量超过拆分大小，会生成多个拆分文件；如果文件数量不超过拆分大小，会生成单个拆分文件包含所有文件）
    （拆分文件的格式根据 --format 参数决定：txt（默认）或 xlsx）

    注意：
    - 汇总文件始终为 txt 格式，不受 --format 参数影响
    - 只有拆分文件才会根据 --format 参数决定格式
    - 这样设计是为了避免汇总文件过大导致 Excel 行数限制问题

    Excel格式说明（仅适用于拆分文件）：
    - Excel文件前两列为"打标签"和"新目录"（供用户填写），第三列为"文件名"（不含扩展名），第四列为"文件类型"（文件扩展名），第五列为"路径"（完整路径）
    - 后续列为路径拆分部分：1级、2级、3级...（根据路径深度自动确定列数）
    - 第一行为表头
    - 例如：路径 d:\\rootdir\\sub1\\sub2\\abc.mp3 会拆分为：
      - 打标签：（空，供用户填写）
      - 新目录：（空，供用户填写）
      - 文件名：abc
      - 文件类型：mp3
      - 路径：d:\\rootdir\\sub1\\sub2\\abc.mp3
      - 1级：d:\\
      - 2级：rootdir
      - 3级：sub1
      - 4级：sub2

使用示例：
    # 处理文件列表，输出到同目录下的"文件列表-3000-utf8"文件夹
    python3 preprocessing/extract-file-paths.py /path/to/文件列表-3000-utf8.txt

    # 指定输出目录
    python3 preprocessing/extract-file-paths.py /path/to/文件列表.txt -o /path/to/output

    # 生成拆分文件（每10000个记录一个文件）
    python3 preprocessing/extract-file-paths.py /path/to/文件列表.txt --split 10000

    # 输出为Excel格式
    python3 preprocessing/extract-file-paths.py /path/to/文件列表.txt --format excel

注意事项：
    1. 输入文件应为UTF-8编码的文本文件，每行一个文件路径
    2. 脚本会自动过滤掉目录路径（无扩展名的行）
    3. 未分类的文件不会导出到任何输出文件
    4. 输出目录如果不存在会自动创建

作者：AI Assistant
创建日期：2025-01-04
最后更新：2025-01-04
"""

import os
import sys
import argparse
from pathlib import Path

try:
    from openpyxl import Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# 定义文件类型分类
# 图片：JPG(JPEG)/png/gif/webp/tiff/raw/svg
IMAGE_EXTENSIONS = {
    'jpg', 'jpeg', 'png', 'gif', 'webp', 'tiff', 'raw', 'svg'
}

# 视频：mp4/mov/m4v/swf/flv/avi/wmv/webm
VIDEO_EXTENSIONS = {
    'mp4', 'mov', 'm4v', 'swf', 'flv', 'avi', 'wmv', 'webm'
}

# 音频：MP3/wav/midi/wma/flac/ape/cda/aac
AUDIO_EXTENSIONS = {
    'mp3', 'wav', 'midi', 'wma', 'flac', 'ape', 'cda', 'aac'
}

# Office文件：Word, Excel, PPT, PDF, 文本文件
OFFICE_EXTENSIONS = {
    # Word
    'doc', 'docx', 'docm', 'dot', 'dotx', 'dotm', 'rtf',
    # Excel
    'xls', 'xlsx', 'xlsm', 'xlt', 'xltx', 'xlsb', 'csv',
    # PPT
    'ppt', 'pptx', 'pot', 'potx', 'pps', 'ppsx',
    # PDF
    'pdf',
    # 文本文件
    'txt', 'md'
}

def get_file_extension(file_path):
    """获取文件扩展名（小写，不含点）"""
    ext = Path(file_path).suffix.lower()
    if ext.startswith('.'):
        ext = ext[1:]
    return ext

def is_file_path(line):
    """判断是否为文件路径（有扩展名）"""
    line = line.strip()
    if not line:
        return False
    
    # 检查是否有扩展名
    ext = get_file_extension(line)
    return bool(ext)

def classify_file_path(file_path):
    """分类文件路径"""
    ext = get_file_extension(file_path)
    
    if ext in IMAGE_EXTENSIONS:
        return 'image'
    elif ext in VIDEO_EXTENSIONS:
        return 'video'
    elif ext in AUDIO_EXTENSIONS:
        return 'audio'
    elif ext in OFFICE_EXTENSIONS:
        return 'office'
    else:
        return None

def write_to_txt(file_path, data_list):
    """写入文本文件"""
    with open(file_path, 'w', encoding='utf-8') as f:
        for line in data_list:
            f.write(line + '\n')

def split_path(path_str):
    """拆分路径为各个部分
    
    Args:
        path_str: 文件路径字符串
        
    Returns:
        list: 路径各部分列表，例如 ['d:\\', 'rootdir', 'sub1', 'sub2', 'abc.mp3']
    """
    parts = []
    
    # 处理 Windows 路径（如 d:\rootdir\sub1\sub2\abc.mp3）
    if '\\' in path_str:
        # Windows 路径，使用反斜杠拆分
        path_parts = path_str.split('\\')
        
        # 处理盘符（如 d:）
        if len(path_parts) > 0 and len(path_parts[0]) == 2 and path_parts[0][1] == ':':
            # 保留盘符和反斜杠作为第一级
            parts.append(path_parts[0] + '\\')
            # 剩余部分
            remaining_parts = path_parts[1:]
        else:
            remaining_parts = path_parts
        
        # 添加剩余部分（过滤空字符串）
        parts.extend([p for p in remaining_parts if p])
    
    # 处理 Unix/Linux/Mac 路径（如 /rootdir/sub1/sub2/abc.mp3）
    elif path_str.startswith('/'):
        path_parts = path_str.split('/')
        # 第一级是根目录
        parts.append('/')
        # 剩余部分（过滤空字符串）
        parts.extend([p for p in path_parts[1:] if p])
    
    # 处理相对路径
    else:
        path_parts = path_str.split('/')
        parts.extend([p for p in path_parts if p])
    
    return parts

def write_to_excel(file_path, data_list):
    """写入Excel文件（前两列为"打标签"和"新目录"，第三列为路径，后续列为路径拆分部分）"""
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl 未安装，无法生成 Excel 文件。请运行：pip3 install openpyxl")
    
    if not data_list:
        return
    
    # 分析所有路径，找出最大深度（只计算目录深度，不包含文件名）
    all_parts = []
    directory_depths = []
    for path in data_list:
        parts = split_path(path)
        all_parts.append(parts)
        # 计算目录深度：如果只有一个部分（如只有盘符），深度为1；否则排除最后一个元素（文件名）
        if len(parts) > 1:
            directory_depth = len(parts) - 1  # 排除文件名
        else:
            directory_depth = len(parts)  # 只有一个部分时保留
        directory_depths.append(directory_depth)
    
    # 找出最大目录深度（不包含文件名）
    max_depth = max(directory_depths) if directory_depths else 0
    
    wb = Workbook()
    ws = wb.active
    ws.title = "文件路径"
    
    # 设置表头：前两列为"打标签"和"新目录"，第三列为"文件名"，第四列为"文件类型"，第五列为"路径"，后续列为路径拆分部分
    headers = ['打标签', '新目录', '文件名', '文件类型', '路径'] + [f'{i}级' for i in range(1, max_depth + 1)]
    for col_idx, header in enumerate(headers, start=1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws[f'{col_letter}1'] = header
    
    # 写入数据
    for row_idx, path in enumerate(data_list, start=2):
        parts = all_parts[row_idx - 2]  # 获取对应的拆分结果
        
        # 只取目录部分，排除文件名（最后一个部分）
        # 如果 parts 只有一个元素（如只有盘符），则保留；否则排除最后一个元素
        if len(parts) > 1:
            directory_parts = parts[:-1]  # 排除最后一个元素（文件名）
        else:
            directory_parts = parts  # 只有一个元素时保留
        
        # 获取文件名（不含扩展名）
        # 手动提取文件名，兼容 Windows 路径格式（反斜杠）
        if '\\' in path:
            # Windows 路径，使用反斜杠拆分
            file_name_with_ext = path.split('\\')[-1]
        elif '/' in path:
            # Unix/Linux/Mac 路径，使用正斜杠拆分
            file_name_with_ext = path.split('/')[-1]
        else:
            # 相对路径或只有文件名
            file_name_with_ext = path
        
        # 去掉扩展名，获取文件名（不含扩展名）
        if '.' in file_name_with_ext:
            # 找到最后一个点，去掉扩展名
            file_name = '.'.join(file_name_with_ext.split('.')[:-1])
        else:
            # 没有扩展名
            file_name = file_name_with_ext
        
        # 获取文件扩展名（文件类型）
        file_ext = get_file_extension(path)
        # 如果扩展名以点开头，去掉点（例如：.jpg -> jpg）
        if file_ext.startswith('.'):
            file_type = file_ext[1:]
        else:
            file_type = file_ext if file_ext else ''
        
        # 写入"打标签"列（第1列，A列）- 留空，供用户填写
        ws[f'A{row_idx}'] = ''
        
        # 写入"新目录"列（第2列，B列）- 留空，供用户填写
        ws[f'B{row_idx}'] = ''
        
        # 写入文件名（第3列，C列）- 不含扩展名的文件名
        ws[f'C{row_idx}'] = file_name
        
        # 写入文件类型（第4列，D列）- 文件扩展名（不含点）
        ws[f'D{row_idx}'] = file_type
        
        # 写入完整路径（第5列，E列）
        ws[f'E{row_idx}'] = path
        
        # 写入路径各部分（第6列及以后，F列及以后）- 只包含目录名，不包含文件名
        for col_idx, part in enumerate(directory_parts, start=6):
            col_letter = ws.cell(row=row_idx, column=col_idx).column_letter
            ws[f'{col_letter}{row_idx}'] = part
    
    # 调整列宽
    ws.column_dimensions['A'].width = 15  # 打标签列
    ws.column_dimensions['B'].width = 30  # 新目录列
    ws.column_dimensions['C'].width = 40  # 文件名列
    ws.column_dimensions['D'].width = 15  # 文件类型列
    ws.column_dimensions['E'].width = 100  # 路径列
    for col_idx in range(6, max_depth + 6):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = 30  # 路径拆分部分列
    
    wb.save(file_path)

def process_file_list(input_file, output_dir, split_size=0, output_format='txt'):
    """处理文件列表并分类导出
    
    Args:
        input_file: 输入文件路径
        output_dir: 输出目录
        split_size: 拆分文件大小（每个文件的最大记录数），0表示不拆分
        output_format: 输出格式，'txt' 或 'excel'，默认为 'txt'
    """
    # 创建输出目录
    os.makedirs(output_dir, exist_ok=True)
    
    # 分类收集文件路径
    categorized_files = {
        'image': [],
        'video': [],
        'audio': [],
        'office': []
    }
    
    # 统计信息
    stats = {
        'image': 0,
        'video': 0,
        'audio': 0,
        'office': 0,
        'other': 0,
        'total_lines': 0,
        'file_paths': 0
    }
    
    # 读取并分类文件
    print(f"正在读取文件：{input_file}")
    with open(input_file, 'r', encoding='utf-8') as f:
        for line_num, line in enumerate(f, 1):
            line = line.strip()
            stats['total_lines'] += 1
            
            # 跳过空行
            if not line:
                continue
            
            # 判断是否为文件路径
            if not is_file_path(line):
                continue
            
            stats['file_paths'] += 1
            
            # 分类文件
            category = classify_file_path(line)
            
            if category:
                categorized_files[category].append(line)
                stats[category] += 1
            else:
                stats['other'] += 1
    
    # 定义输出文件名模板
    file_templates = {
        'image': '图片文件路径',
        'video': '视频文件路径',
        'audio': '音频文件路径',
        'office': 'Office文件路径'
    }
    
    # 汇总文件始终使用 txt 格式
    # 拆分文件根据 output_format 参数决定格式
    if output_format == 'excel':
        split_file_ext = '.xlsx'
        split_write_func = write_to_excel
    else:
        split_file_ext = '.txt'
        split_write_func = write_to_txt
    
    # 生成带数量的文件名并写入文件
    output_files = {}
    split_files = []
    
    for category, template in file_templates.items():
        count = stats[category]
        if count > 0:
            files_list = categorized_files[category]
            
            # 生成汇总文件（始终使用 txt 格式）
            filename = f"{template}-共{count}个.txt"
            file_path = os.path.join(output_dir, filename)
            output_files[category] = file_path
            
            # 写入汇总文件（始终使用 txt 格式）
            try:
                write_to_txt(file_path, files_list)
            except ImportError as e:
                print(f"\n错误：{e}")
                sys.exit(1)
            
            # 如果需要拆分，生成拆分文件（根据 output_format 决定格式）
            if split_size > 0:
                if count > split_size:
                    # 文件数量超过拆分大小，生成多个拆分文件
                    num_splits = (count + split_size - 1) // split_size  # 向上取整
                    for i in range(num_splits):
                        start_idx = i * split_size
                        end_idx = min((i + 1) * split_size, count)
                        start_num = start_idx + 1  # 从1开始计数
                        end_num = end_idx
                        
                        # 生成拆分文件名（根据 output_format 决定扩展名）
                        split_filename = f"{template}-第{start_num}-{end_num}个{split_file_ext}"
                        split_file_path = os.path.join(output_dir, split_filename)
                        split_files.append(split_file_path)
                        
                        # 写入拆分文件（根据 output_format 决定格式）
                        try:
                            split_write_func(split_file_path, files_list[start_idx:end_idx])
                        except ImportError as e:
                            print(f"\n错误：{e}")
                            sys.exit(1)
                else:
                    # 文件数量不超过拆分大小，生成单个拆分文件（包含所有文件）
                    split_filename = f"{template}-第1-{count}个{split_file_ext}"
                    split_file_path = os.path.join(output_dir, split_filename)
                    split_files.append(split_file_path)
                    
                    # 写入拆分文件（根据 output_format 决定格式）
                    try:
                        split_write_func(split_file_path, files_list)
                    except ImportError as e:
                        print(f"\n错误：{e}")
                        sys.exit(1)
    
    # 打印统计信息
    print("\n" + "="*60)
    print("处理完成！统计信息：")
    print("="*60)
    print(f"总行数：{stats['total_lines']:,}")
    print(f"文件路径数：{stats['file_paths']:,}")
    print(f"\n分类统计：")
    print(f"  图片文件：{stats['image']:,} 个")
    print(f"  视频文件：{stats['video']:,} 个")
    print(f"  音频文件：{stats['audio']:,} 个")
    print(f"  Office文件：{stats['office']:,} 个")
    print(f"  其他文件：{stats['other']:,} 个")
    print("\n输出文件：")
    for category, file_path in output_files.items():
        print(f"  {os.path.basename(file_path)}: {stats[category]:,} 个文件")
    
    if split_files:
        print(f"\n拆分文件（每{split_size:,}个记录一个文件）：")
        for split_file in split_files:
            print(f"  {os.path.basename(split_file)}")
    
    print("="*60)

def main():
    parser = argparse.ArgumentParser(
        description='文件路径分类导出脚本：读取文件列表，按类型分类导出：图片、视频、音频、Office文件',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
使用示例：
  # 基本用法（输出目录自动使用输入文件名）
  python3 preprocessing/extract-file-paths.py /path/to/文件列表.txt

  # 指定输出目录
  python3 preprocessing/extract-file-paths.py /path/to/文件列表.txt -o /path/to/output

  # 生成拆分文件（每10000个记录一个文件）
  python3 preprocessing/extract-file-paths.py /path/to/文件列表.txt --split 10000

  # 输出为Excel格式
  python3 preprocessing/extract-file-paths.py /path/to/文件列表.txt --format excel

输出文件：
  脚本会在输出目录中生成以下文件（文件名包含记录数）：
  - 图片文件路径-共N个.txt    - 包含所有图片文件路径（汇总文件，始终为txt格式）
  - 视频文件路径-共N个.txt    - 包含所有视频文件路径（汇总文件，始终为txt格式）
  - 音频文件路径-共N个.txt    - 包含所有音频文件路径（汇总文件，始终为txt格式）
  - Office文件路径-共N个.txt  - 包含所有Office和文本文件路径（汇总文件，始终为txt格式）
  （N为实际文件数量，如果某类文件数量为0，则不会生成对应文件）

  如果指定了 --split 参数，还会生成拆分文件：
  - 图片文件路径-第1-10000个.{txt|xlsx}    - 第1到10000个记录
  - 图片文件路径-第10001-20000个.{txt|xlsx} - 第10001到20000个记录
  （如果文件数量超过拆分大小，会生成多个拆分文件；如果文件数量不超过拆分大小，会生成单个拆分文件包含所有文件）
  （拆分文件的格式根据 --format 参数决定：txt（默认）或 xlsx）

  注意：
  - 汇总文件始终为 txt 格式，不受 --format 参数影响
  - 只有拆分文件才会根据 --format 参数决定格式
  - 这样设计是为了避免汇总文件过大导致 Excel 行数限制问题

  Excel格式说明（仅适用于拆分文件）：
  - Excel文件前两列为"打标签"和"新目录"（供用户填写），第三列为"文件名"（不含扩展名），第四列为"文件类型"（文件扩展名），第五列为"路径"（完整路径）
  - 后续列为路径拆分部分：1级、2级、3级...（根据路径深度自动确定列数）
  - 第一行为表头
  - 例如：路径 d:\\rootdir\\sub1\\sub2\\abc.mp3 会拆分为：
    - 打标签：（空，供用户填写）
    - 新目录：（空，供用户填写）
    - 文件名：abc
    - 文件类型：mp3
    - 路径：d:\\rootdir\\sub1\\sub2\\abc.mp3
    - 1级：d:\\
    - 2级：rootdir
    - 3级：sub1
    - 4级：sub2
        """
    )
    parser.add_argument(
        'input_file',
        type=str,
        help='输入文件路径（文件列表）'
    )
    parser.add_argument(
        '-o', '--output',
        type=str,
        default=None,
        help='输出目录（可选，默认使用输入文件名）'
    )
    parser.add_argument(
        '-s', '--split',
        type=int,
        default=0,
        metavar='SIZE',
        help='拆分文件大小（每个文件的最大记录数），如果指定此参数，会在汇总文件基础上生成拆分文件。例如：--split 10000 会生成每10000个记录一个文件'
    )
    parser.add_argument(
        '-f', '--format',
        type=str,
        choices=['txt', 'excel'],
        default='txt',
        help='拆分文件的输出格式：txt（文本文件，默认）或 excel（Excel文件）。注意：汇总文件始终为 txt 格式'
    )
    
    args = parser.parse_args()
    
    input_file = args.input_file
    
    # 检查输入文件是否存在
    if not os.path.exists(input_file):
        print(f"错误：输入文件不存在：{input_file}")
        sys.exit(1)
    
    # 生成输出目录名称
    if args.output:
        output_dir = args.output
    else:
        # 使用输入文件名（不含扩展名）作为输出目录名
        input_path = Path(input_file)
        output_dir_name = input_path.stem  # 不含扩展名的文件名
        # 输出目录放在输入文件所在目录
        output_dir = str(input_path.parent / output_dir_name)
    
    # 验证拆分参数
    split_size = args.split
    if split_size < 0:
        print(f"错误：拆分大小必须大于0，当前值：{split_size}")
        sys.exit(1)
    
    # 验证输出格式
    output_format = args.format.lower()
    if output_format == 'excel' and not OPENPYXL_AVAILABLE:
        print("错误：输出格式为 excel 但 openpyxl 未安装。")
        print("请运行：pip3 install openpyxl")
        sys.exit(1)
    
    # 创建输出目录
    os.makedirs(output_dir, exist_ok=True)
    
    process_file_list(input_file, output_dir, split_size, output_format)
    print(f"\n✅ 所有文件已导出到：{output_dir}")

if __name__ == '__main__':
    main()

