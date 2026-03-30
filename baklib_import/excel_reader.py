#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 文件读取模块

功能：
- 从 Excel 文件读取文件路径列表
- 管理导入状态和 DAM ID
- 支持备份和恢复

维护：Baklib Tools
创建日期：2026-01-06
"""

import os
import logging
from typing import List, Dict, Optional, Tuple
from datetime import datetime

try:
    import openpyxl
    from openpyxl import styles
    from openpyxl import utils
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def _shift_column_letter(column: str, delta: int) -> str:
    """将列字母按 delta 右移（delta 可为负）。"""
    idx = utils.column_index_from_string(column.strip().upper()) + delta
    if idx < 1:
        raise ValueError(f"列偏移无效: {column!r} + {delta}")
    return utils.get_column_letter(idx)


class ExcelReader:
    """Excel 文件读取器"""
    
    def __init__(self, excel_path: str):
        """
        初始化 Excel 读取器
        
        Args:
            excel_path: Excel 文件路径
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl 未安装，请运行：pip install openpyxl")
        
        if not os.path.exists(excel_path):
            raise FileNotFoundError(f"Excel 文件不存在：{excel_path}")
        
        self.excel_path = excel_path
        self.wb = None
        self.ws = None
        self._load_workbook()
    
    def _load_workbook(self):
        """加载工作簿"""
        self.wb = openpyxl.load_workbook(self.excel_path)
        self.ws = self.wb.active
    
    def ensure_status_columns(self, status_column: str = None, dam_id_column: str = None) -> Tuple[str, str, int]:
        """
        确保导入状态列和DAM ID列存在
        
        Args:
            status_column: 导入状态列（如果为None，先查找现有列，找不到则自动添加）
            dam_id_column: DAM ID列（如果为None，先查找现有列，找不到则自动添加）
        
        Returns:
            (status_column, dam_id_column, left_columns_inserted)
            left_columns_inserted: 本次在表头左侧新插入的列数（0 或 2），用于解析路径列偏移
        """
        left_inserted = 0

        # 如果未指定列，先查找是否已存在；不存在则在**最左侧**插入两列（便于深目录时查看状态）
        if status_column is None:
            status_column = self._find_column_by_header('导入状态')
            if status_column is None:
                self.ws.insert_cols(1, 2)
                left_inserted = 2
                status_column = 'A'
                dam_id_column = 'B'
                sh = self.ws['A1']
                dh = self.ws['B1']
                sh.value = '导入状态'
                sh.font = styles.Font(bold=True)
                dh.value = 'DAM ID'
                dh.font = styles.Font(bold=True)
                return status_column, dam_id_column, left_inserted

        if dam_id_column is None:
            dam_id_column = self._find_column_by_header('DAM ID')
            if dam_id_column is None:
                status_col_num = utils.column_index_from_string(status_column)
                dam_id_col_num = status_col_num + 1
                dam_id_column = utils.get_column_letter(dam_id_col_num)

        # 设置标题行（第1行）
        status_header = self.ws[f'{status_column}1']
        dam_id_header = self.ws[f'{dam_id_column}1']

        if not status_header.value or str(status_header.value).strip() != '导入状态':
            status_header.value = '导入状态'
            status_header.font = styles.Font(bold=True)

        if not dam_id_header.value or str(dam_id_header.value).strip() != 'DAM ID':
            dam_id_header.value = 'DAM ID'
            dam_id_header.font = styles.Font(bold=True)

        return status_column, dam_id_column, left_inserted
    
    def _find_column_by_header(self, header_text: str) -> Optional[str]:
        """
        根据标题文本查找列
        
        Args:
            header_text: 标题文本
        
        Returns:
            列名（如 'A', 'B'），如果未找到返回 None
        """
        for col_idx in range(1, self.ws.max_column + 1):
            cell = self.ws.cell(row=1, column=col_idx)
            if cell.value and str(cell.value).strip() == header_text:
                return utils.get_column_letter(col_idx)
        return None
    
    def read_file_list(self, start_row: int = 2, path_column: str = 'E',
                      status_column: str = None, dam_id_column: str = None,
                      max_rows: int = None) -> List[Dict]:
        """
        读取文件列表
        
        Args:
            start_row: 开始行（默认第2行，第1行是标题）
            path_column: 文件路径列（默认 E 列）
            status_column: 导入状态列（如果为None，自动查找或添加）
            dam_id_column: DAM ID列（如果为None，自动查找或添加）
            max_rows: 最大读取行数（可选，用于测试）
        
        Returns:
            文件列表，每个元素包含：
            - row_idx: 行号
            - file_path: 文件路径
            - status: 当前状态（如果已存在）
            - dam_id: DAM ID（如果已存在）
        """
        # 确保状态列存在
        status_column, dam_id_column, left_inserted = self.ensure_status_columns(
            status_column, dam_id_column
        )

        # 路径列：优先按表头「路径」定位（插入状态列后原 E 列会右移，表头仍可靠）
        path_col_resolved = self._find_column_by_header('路径')
        if path_col_resolved is None:
            path_col_resolved = (
                _shift_column_letter(path_column, left_inserted)
                if left_inserted
                else path_column
            )
        if left_inserted:
            logging.info(
                "已在工作表最左侧插入「导入状态」「DAM ID」列；"
                "文件路径列使用 %s（配置为 %s）",
                path_col_resolved,
                path_column,
            )

        file_list = []
        processed_count = 0
        
        for row_idx in range(start_row, self.ws.max_row + 1):
            # 检查是否达到最大行数限制
            if max_rows and processed_count >= max_rows:
                logging.info(f"已达到最大读取行数限制（{max_rows}），停止读取")
                break
            
            # 读取文件路径
            path_cell = self.ws[f'{path_col_resolved}{row_idx}']
            if not path_cell.value:
                continue
            
            file_path = str(path_cell.value).strip()
            if not file_path:
                continue
            
            # 读取状态和 DAM ID
            status_cell = self.ws[f'{status_column}{row_idx}']
            dam_id_cell = self.ws[f'{dam_id_column}{row_idx}']
            
            status = str(status_cell.value).strip() if status_cell.value else ''
            dam_id = str(dam_id_cell.value).strip() if dam_id_cell.value else ''
            
            # 如果状态是"成功"，跳过
            if status == '成功':
                logging.debug(f"[{row_idx}] 跳过（已导入）：{file_path}")
                continue
            
            file_list.append({
                'row_idx': row_idx,
                'file_path': file_path,
                'status': status,
                'dam_id': dam_id
            })
            processed_count += 1
        
        return file_list
    
    def update_status(self, row_idx: int, status: str, dam_id: str = '',
                     status_column: str = None, dam_id_column: str = None):
        """
        更新导入状态
        
        Args:
            row_idx: 行号
            status: 状态（如"成功"、"失败"、"导入中"）
            dam_id: DAM ID（可选）
            status_column: 状态列（如果为None，自动查找）
            dam_id_column: DAM ID列（如果为None，自动查找）
        """
        if status_column is None:
            status_column = self._find_column_by_header('导入状态')
        if dam_id_column is None:
            dam_id_column = self._find_column_by_header('DAM ID')
        
        if status_column:
            status_cell = self.ws[f'{status_column}{row_idx}']
            status_cell.value = status
            # 设置颜色
            if status == '成功':
                status_cell.fill = styles.PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            elif status.startswith('失败'):
                status_cell.fill = styles.PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        
        if dam_id_column and dam_id:
            dam_id_cell = self.ws[f'{dam_id_column}{row_idx}']
            dam_id_cell.value = dam_id
    
    def save(self):
        """保存 Excel 文件（使用临时文件确保写入安全）
        
        说明：
        - 先写入同目录下的临时文件
        - 写入成功后再使用 os.replace 原子性地替换原文件
        - 如果在写入过程中被 Ctrl+C 中断，原始 Excel 文件不会被修改
        """
        tmp_path = self.excel_path + ".tmp"
        try:
            # 先保存到临时文件，避免直接覆盖原文件
            self.wb.save(tmp_path)
            # 写入成功后再原子性替换原文件
            os.replace(tmp_path, self.excel_path)
        finally:
            # 清理可能残留的临时文件（正常情况下这里不存在）
            if os.path.exists(tmp_path):
                try:
                    os.remove(tmp_path)
                except OSError:
                    # 清理失败不影响主流程
                    pass
    
    def create_backup(self) -> str:
        """
        创建备份文件
        
        Returns:
            备份文件路径
        """
        excel_dir = os.path.dirname(os.path.abspath(self.excel_path))
        backup_dir = os.path.join(excel_dir, "backup")
        os.makedirs(backup_dir, exist_ok=True)

        base_name = os.path.basename(self.excel_path)
        stem, ext = os.path.splitext(base_name)
        if not ext:
            ext = ".xlsx"

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_file_name = f"{stem}_backup_{ts}{ext}"
        backup_path = os.path.join(backup_dir, backup_file_name)

        self.wb.save(backup_path)
        logging.info(f"已创建备份文件：{backup_path}")
        return backup_path
    
    def close(self):
        """关闭工作簿"""
        if self.wb:
            self.wb.close()
    
    def __enter__(self):
        """上下文管理器入口"""
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        """上下文管理器出口"""
        self.close()

